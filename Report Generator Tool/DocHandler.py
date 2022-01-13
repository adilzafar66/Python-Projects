#------------------------------------------------------------------------------------------#
# Author: Adil Zafar Khan
# Last Edit Date: 12/22/2021
# Description:
"""
    DocHandler contains the DataImporter class which is responsible for handling all the
    word document processes. It imports excel tables, finds and replaces words, imports SLD
    diagrams and creates a TCC graphs table.
"""
#------------------------------------------------------------------------------------------#

#Import all the required libraries
import re
import os
from pathlib import Path
import comtypes.client
import pythoncom
from win32com import client
from PIL import Image
from pdf2image import convert_from_path
from pytesseract import pytesseract
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#Set the path of OCR to the exisiting exe file in C Drive
path_to_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.tesseract_cmd = path_to_tesseract

#Data Importer class does the following:
#(1) Inserts tables into the word document
#(2) Searches and replaces keywords with input words
#(3) Generates a TCC table into the word document
#(4) Converts the final docx file to pdf

#Note: the class uses different libraries to handle excel files simaltaneously namely openpyxl and pywin32

class DataImporter():

    #Initilize class
    #Set docPath as the path for the report template
    def __init__(self, docPath):

        docPath = Path(docPath)
        pythoncom.CoInitialize()
        
        #Open word in python with pywin32
        self.word = client.gencache.EnsureDispatch("Word.Application")
        self.doc = self.word.Documents.Open(str(docPath))
        self.docPath = docPath

        #Open excel in python with pywin32
        self.excel = client.gencache.EnsureDispatch("Excel.Application")

    def __del__(self):

        #Delete all the word variables
        self.doc.Close(SaveChanges = 0)
        self.word.Quit()
        del self.word
        
        #Delete all the excel variables
        self.closeExcelWbs()
        self.excel.Quit()
        del self.excel

    #Returns the position of the first row that is empty in an excel sheet
    def getFirstEmptyRowNum(self, sheet):

        #Loop through the first cell in each row
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value == None:
                    return cell.row

        return sheet.max_row + 1
    
    #Gets the maximum filled columns range (string) in an excel sheet
    def __getColRange(self, wbPath, sheetNum):

        #Use openpyxl to get max columns
        wb = load_workbook(wbPath)
        ws = wb.worksheets[sheetNum]
        colRange = 'A1:' + get_column_letter(ws.max_column) + str(self.getFirstEmptyRowNum(ws) - 1)
        wb.close()
        
        return colRange

    def findText(self, text):

        #Move the cursor to the start of the document
        #wdStory = 6
        selectWord = self.word.Selection
        selectWord.HomeKey(6)
        
        #Find the afterText in the word document and select it
        selectWord.Find.ClearFormatting()
        selectWord.Find.Forward = True
        selectWord.Find.MatchWholeWord = True
        selectWord.Find.MatchCase = True
        selectWord.Find.Wrap = 1
        selectWord.Find.Text = text

        return selectWord
        
    #Imports a table sheet from an excel file to the destination report at docPath
    def importExcelTable(self, wbPath, sheetNum, numHeaderRows = 1, afterTextDict = None, close = True):

        wb = self.excel.Workbooks.Open(wbPath)
        ws = wb.Worksheets(sheetNum)

        #xlToLeft = -4159
        
        #Get the max number of filled cells range in a sheet
        #lCol = ws.Cells(2, self.excel.Columns.Count).End(xlToLeft).Column
        colRange = self.__getColRange(wbPath, sheetNum - 1)

        #Find the afterText in the word document and select it
        selectWord = self.findText(Path(wbPath).stem)
        selectWord.Find.Execute()

        if not selectWord.Find.Found:
            
            if close:
                #Close the excel file without saving
                wb.Close(SaveChanges = 0)
                
            return -1

        #Copy all the contents in the colRange range
        ws.Range(colRange).Copy()
        
        #Paste excel table at the cursor location
        selectWord.PasteExcelTable(False, False, False)

        #Format the table
        selectWord.Tables(1).Rows.Alignment = 1
        selectWord.Tables(1).AutoFitBehavior(2)
        selectWord.Tables(1).AllowAutoFit = True
        selectWord.Tables(1).Range.ParagraphFormat.SpaceAfter = 0

        for i in range(1, numHeaderRows + 1):
            selectWord.Tables(1).Rows(i).HeadingFormat = True

        #Move cursor to the start of the document
        #wdStory = 6
        selectWord.HomeKey(6)

        #Empty the clipboard
        self.excel.CutCopyMode = False

        if close:
            #Close the excel file without saving
            wb.Close(SaveChanges = 0)

    def importImage(self, imgPath, afterText):

        #Find the afterText in the word document and select it
        selectWord = self.findText(Path(imgPath).stem)
        selectWord.Find.Execute()

        if not selectWord.Find.Found:     
            return -1

        imgAdded = selectWord.InlineShapes.AddPicture(FileName = imgPath, LinkToFile=False, SaveWithDocument=True)
        imgAdded.ScaleHeight = 10
        imgAdded.ScaleWidth = 10

    #Finds and replaces all the instances of a text in the word document with a passed string
    def findReplaceAll(self, findText, replaceText):

        #Don't do anything if the replaceText is an empty string
        if replaceText == '':
            return
        
        selectWord = self.findText(findText)
        selectWord.Find.Replacement.ClearFormatting()
        selectWord.Find.Replacement.Text = replaceText
        
        #wdReplaceAll = 2
        selectWord.Find.Execute(Replace = 2)
        
        #Search in headers and footers
        for i in range(1, self.doc.Sections.Count + 1):

            try:
                self.doc.Sections(1).Headers(i).Range.Find.Execute(FindText = findText, Format = False,
                                                                   ReplaceWith = replaceText, Replace = 2,
                                                                   Wrap = 1, MatchWholeWord = True,
                                                                   MatchCase = True)
            except:
                pass
            
            try:
                self.doc.Sections(1).Footers(i).Range.Find.Execute(FindText = findText, Format = False,
                                                                   ReplaceWith = replaceText, Replace = 2,
                                                                   Wrap = 1, MatchWholeWord = True,
                                                                   MatchCase = True)
            except:
                pass

    # Finds and deletes all instances of texts in findTextList from the document
    def findDeleteAll(self, findTextList):

        for findText in findTextList:

            # Initialize selection
            selectWord = self.findText(findText)
            selectWord.Find.Execute()

            # Loop until all the instances are deleted
            # Delete the new line as well
            while(selectWord.Find.Found):
                selectWord.Delete()
                selectWord.Delete()
                selectWord.Find.Execute()

    #Creates a map of texts that need to be replaced with their replacement texts
    #Calls findReplaceAll for each pair
    def insertWords(self, textList, replacementList):

        #Make the list sizes equal if they are not equal
        if len(textList) != len(replacementList):
            replacementList[: len(textList) - 1]

        #Call findReplaceAll for each item in the paired list
        for i in range(len(textList)):
            self.findReplaceAll(textList[i], replacementList[i])

    #Runs an excel macro in an excel file
    def runXlMacro(self, wbPath, macroName, saveCopy = False, savePath = None):

        wb = self.excel.Workbooks.Open(wbPath, ReadOnly = 1)
        self.excel.Application.Run(Path(wbPath).name + macroName)

        #Save a copy of filled sheet in the specified path
        if saveCopy:
            wb.SaveCopyAs(str(savePath))

        wb.Close(SaveChanges = 0)

    # Returns a list of sheet names that are not empty in an excel workbook
    def sheetsFilledList(self, wbPath):

        sheetList = []
        wb = self.excel.Workbooks.Open(wbPath, ReadOnly = 1)
        
        # Loop through all the sheets starting from sheet 3
        for i in range(3, wb.Worksheets.Count + 1):
            if wb.Worksheets(i).Cells(4, 1).Value != None:
                sheetList.append(i)

        wb.Close(SaveChanges = 0)
        
        return sheetList
    
    #Use regex to find the graph name in title
    def __findNameTitle(self, text):

        return re.findall("(?<=" + self.__findNumTitle(text)[0] + ").*?(?=[.])", text)

    #Use regex to find the graph number in title
    def __findNumTitle(self, text):

        return re.findall(r"(?:[0-9][0-9]?|100)(?:[.][0-9][0-9]?|100?)?(?:[.][0-9][0-9]?|100)-", text)

    #Use regex to find the fault type of the graph
    def __findFault(self, text):

        return re.findall("(?<=Fault: ).*?(?=\n)", text)

    #Creates a list of all the TCC graph titles splitted into three columns
    #Col 1 = Number, Col 2 = Name, Col 3 = Fault
    def createTCCTitleList(self, pdfPath):

        self.titleTable = []

        folderPath = Path(pdfPath).parent

        #Loop through all the files in the folderPath directory
        for filename in os.listdir(folderPath):

            #If it is a jpeg file
            if filename.endswith(".pdf") and filename != Path(pdfPath).name:

                filePath = Path(folderPath, filename)
                
                #Convert the pdf file to image
                image = convert_from_path(filePath, 500)[0]
                
                #Scan the image for text
                text = pytesseract.image_to_string(image)[:-1]

                #Close image
                image.close()
                
                #Get the relevant information from the extracted text
                try:
                    fault = self.__findFault(text)[0]
                except:
                    fault = 'None'

                try:
                    numTitle = self.__findNumTitle(text)[0]
                except:
                    numTitle = 'None'

                try:
                    nameTitle = self.__findNameTitle(text)[0]
                except:
                    nameTitle = 'None'

                #Format the information before storing it into a list
                numTitle = numTitle[ : numTitle.find('-')]

                #Store the information in the list
                self.titleTable.append([numTitle, nameTitle, fault])

    #Creates a TCC table in the word document using the titleList 
    def createTCCTable(self, titleList, findText):

        #titleList is a 2D list

        #Find the insertion location for the table

        insertionPlace = self.findText(findText)
        insertionPlace.Find.Execute()

        #Create a new table
        TCCTable = self.doc.Tables.Add(Range = insertionPlace.Range, NumRows = len(titleList) + 1, NumColumns = 3, AutoFitBehavior = 2)

        #Format the table
        TCCTable.Borders.InsideLineStyle = 1 
        TCCTable.Borders.OutsideLineStyle = 1
        TCCTable.Rows.SetHeight(RowHeight = 0.21 * 72, HeightRule = 2)
        TCCTable.Columns(1).Width = 0.57 * 72
        TCCTable.Columns(2).Width = 2.14 * 72
        TCCTable.Columns(3).Width = 4.17 * 72

        #Remove default formatting
        TCCTable.ApplyStyleFirstColumn = False 
        TCCTable.ApplyStyleHeadingRows = False 
        TCCTable.ApplyStyleLastColumn = False 
        TCCTable.ApplyStyleLastRow = False

        #Insert data into header cells
        TCCTable.Rows(1).Shading.BackgroundPatternColor = 15921906
        TCCTable.Cell(1,1).Range.InsertAfter("Graph:")
        TCCTable.Cell(1,2).Range.InsertAfter("Title:")
        TCCTable.Cell(1,3).Range.InsertAfter("Description:")

        #Insert data into the body using the titleList
        for row in range(len(titleList)):
            TCCTable.Cell(row + 2,1).Range.InsertAfter(titleList[row][0])

            #If there is only one element in a row then it must be another heading
            try:
                TCCTable.Cell(row + 2,2).Range.InsertAfter(titleList[row][1])
            except:
                #Merge the and centre align the cells because it is a heading
                TCCTable.Cell(row + 2, 1).Merge(TCCTable.Cell(row + 2, 3))
                TCCTable.Cell(row + 2, 1).Range.ParagraphFormat.Alignment = 1

        # Sort the table with the first column 
        TCCTable.Sort(ExcludeHeader = True, FieldNumber = 1)

    #Creates a BCH table in word by scanning the BCH info document
    def createBchTable(self):

        pass

    #Save the word document as PDF at the same location
    def convertToPdf(self, path = None):
        
        wdFormatPDF = 17
        
        if path == None:
            path = str(Path(self.docPath).with_suffix('.pdf'))
        
        self.doc.SaveAs(str(path), FileFormat = wdFormatPDF)

    def closeExcelWbs(self):

        try:
            
            for i in range(1, self.excel.Workbooks.Count + 1):
                self.excel.Workbooks(i).Close(SaveChanges = 0)
        except:
            pass

    #Save the word document
    def save(self):
        self.doc.Save()

    #Save a copy of the doc to the location specified
    def saveAs(self, path):

        self.doc.SaveAs2(FileName = str(path))
