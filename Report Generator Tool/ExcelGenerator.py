#------------------------------------------------------------------------------------------#
# Author: Adil Zafar Khan
# Last Edit Date: 12/22/2021
# Description:
"""
    ExcelGenerator is a module that contains classes responsible for generating excel sheets
    to be later imported to the word document. The output excel sheets are saved in the excel outputs
    directory which is later accessed by the DataImporter class.
"""
#------------------------------------------------------------------------------------------#

#Import the require libraries
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
import win32com.client as win32
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from ExcelHandler import ReportsTableHandler

#Class that generates an INT Report
class IntReport(ReportsTableHandler):

    #Function responsible for generating a single INT Report from a directory path
    #Using Regex to filter out the files we do not need
    #Make sure that the file naming convention is followed as stated in the User Guide
    
    def generateTable(self, dirPath, outputDir = os.getcwd(),
                      outputName = "INT_Report.xlsx", regexName = "INT_.*_.*_.*"):

        #------------------------INITIALIZING DEFAULT VARIABLES--------------------------#
        
        #Set the length of header columns and body columns
        hNumCols = 4
        bNumCols = 1

        #Set the first row and column to be extracted from the source file
        sRow = 2
        sCol = "^Int. Adj. Symm. kA$|^Ib Sym$"

        #Flag to check if data was added to the IntReport class worksheet
        dataAdded = False

        #------------------------------CREATING HEADERS---------------------------------#
        
        #Create a file names list for files
        fileSet = self.createFileSet(dirPath, "(?<=_)....?(?=[.])|(?<=_)...?(?=[.])")

        #Specify the order of data columns with file names
        order = ["PRES", "GEN", "ULT", "CTT"]

        #Unordered header list generated from the filename **_TYPE_**
        unorderedHeaders = self.generateHeaders(dirPath, regexName)

        #Ordered header list generated from fixed order list above
        orderedHeaders = self.listUnion(order, unorderedHeaders)

        #Remainder headers whose order does not matter
        remHeaders = self.listDiff(unorderedHeaders, orderedHeaders)

        #All ordered header list
        allOrderedHeaders = orderedHeaders + remHeaders

        #-------------------------------CREATING BODY-----------------------------------#
        
        #For the first group, set the destination row as three since we have two header rows
        dRow = 3

        #Loop through all the groups such as "LVCB" or "FUSE"
        for group in fileSet:
            
            fileNames = self.createFileList(dirPath, regexName + "_" + group)

            #If no files are found for this group, skip the iteration
            if len(fileNames) == 0:
                continue

            #If the iteration is not skipped, reorder the fileNames list
            fileNamesOrdered = self.fileListReorder(fileNames, order)

            #Run the data extraction loop. addClosingCols is set to True since we need the device capability columns
            self.loopData(dirPath, fileNamesOrdered, dRow, 1, sRow, sCol, hNumCols,
                          bNumCols, addClosingCols = True)

            #Set dataAdded as True so that the file is generated at the end
            dataAdded = True

            #Set the destination row as two for the next itertion since the next iteration has no header rows
            dRow = 2

        #If no data is added, exit function
        if not dataAdded:
            return

        #-----------------------------FORMATTING TABLE---------------------------------#
        
        #Move the fourth column two steps to the left
        self.moveColumn(4, -2)

        #Format the sheet
        self.formatTable(hNumCols, bNumCols)

        #------------------------------ADDING HEADERS--------------------------------#

        #Create a headers and ubheaders list
        headers = ["Device"] + ["Interrupting Duty (Adjusted Symm. kA)"] + ["Device Capability"]
        subheaders = ["ID", "Voltage", "Bus" ,"Device"] + allOrderedHeaders + ["Rating Int. kA"]

        #Set the length of the merged header cells for the header columns and the body columns
        hMergeLen = 4
        bMergeLen = int(((self.ws.max_column - hNumCols)/bNumCols) - 3)

        #Merge header cells and add headers
        self.mergeHeaderCells(hMergeLen, bMergeLen)
        self.addHeaders(hNumCols, bNumCols, bMergeLen , headers, subheaders)

        # Create a 2D list of columns to compare
        # The reference column comes first
        cmpCols = [["Rating Int. kA", item] for item in allOrderedHeaders]

        # Color fill the cells that do not meet the conditions in the function
        self.fillCmpCells(2, cmpCols)

        #---------------------------------SAVING------------------------------------#
        
        #Change the filename to some tag
        self.saveWorkbook(dirPath = outputDir, fileName = outputName)

        #Autofit columns
        self.autofitColumns(self.wbPath)

#Class that generates a MOM Report
class MomReport(ReportsTableHandler):

    #Function responsible for generating a single INT Report from a directory path
    #Using Regex to filter out the files we do not need
    #Make sure that the file naming convention is followed as stated in the User Guide
    
    def generateTable(self, dirPath, outputDir = os.getcwd(),
                      outputName = "MOM_Report.xlsx", regexName = "MOM_.*_.*_.*"):

        #------------------------INITIALIZING DEFAULT VARIABLES--------------------------#
        
        #Set the length of header columns and body columns
        hNumCols = 3
        bNumCols = 2

        sRow = 2
        sCol = "^Symm. kA$"

        #Create a file names list for files in the specified directory
        fileNames = self.createFileList(dirPath, regexName)

        if len(fileNames) == 0:
            return

        #------------------------------CREATING HEADERS---------------------------------#
        
        #Specify the order of data columns with file names
        order = ["PRES", "GEN", "ULT", "CTT"]

        #All unordered header list generated from the filename **_TYPE_**
        unorderedHeaders = self.generateHeaders(dirPath, regexName)

        #Orderable ordered header list generated from fixed order list above
        orderedHeaders = self.listUnion(order, unorderedHeaders)

        #Remainder headers whose order does not matter
        remHeaders = self.listDiff(unorderedHeaders, orderedHeaders)

        #All ordered header list
        allOrderedHeaders = orderedHeaders + remHeaders

        #Reorder file names by using the order list
        fileNamesOrdered = self.fileListReorder(fileNames, order)

        #-------------------------------CREATING BODY-----------------------------------#
        
        #Loop through all the files in the reordered file list and generate excel files with its data
        self.loopData(dirPath, fileNamesOrdered, 3, 1, sRow, sCol, hNumCols,
                      bNumCols, addClosingCols = True)

        #-----------------------------FORMATTING TABLE---------------------------------#
        
        #Delete rows with 'Bus' in the ID Name
        self.deleteBusRows()
        
        #Format the sheet
        self.formatTable(hNumCols, bNumCols)

        #------------------------------ADDING HEADERS--------------------------------#
        
        #Create a headers and subheaders list
        headers = ["Device"] + ["Momentary Duty: " + s for s in allOrderedHeaders] + ["Device Capability"]
        subheaders = ["ID", "Voltage", "Device"] + ["Symm. kA rms", "Asymm. kA rms"] * int(((self.ws.max_column - hNumCols)/bNumCols) - 1)

        #Set the length of the merged header cells for the header columns and the body columns
        hMergeLen = hNumCols
        bMergeLen = bNumCols

        #Merge header cells and add headers
        self.mergeHeaderCells(hMergeLen, bMergeLen)
        self.addHeaders(hNumCols, bNumCols, bMergeLen , headers, subheaders)

        for i in range(0, len(allOrderedHeaders) * 3 - 1, 3):

            symCmpCol = self.getColumnByHeaderName(self.ws, "Symm. kA rms", 2) + i 
            asymCmpCol = self.getColumnByHeaderName(self.ws, "Asymm. kA rms", 2) + i

            cmpCols = [[self.ws.max_column - 1, symCmpCol]] * len(allOrderedHeaders)
            self.fillCmpCells(2, cmpCols)

            cmpCols = [[self.ws.max_column, asymCmpCol]] * len(allOrderedHeaders)
            self.fillCmpCells(2, cmpCols)

        #---------------------------------SAVING------------------------------------#
        
        #Change the filename to some tag
        self.saveWorkbook(dirPath = outputDir, fileName = outputName)

        #Autofit columns
        self.autofitColumns(self.wbPath)

    #Function responsible to delete all the rows who's ID starts with the word 'bus'
    def deleteBusRows(self):

        #Start from the second row
        row = 2

        #Loop until the last row
        while(row <= self.ws.max_row):

            #Captalization does not matter since we are using lower() function
            if str(self.ws.cell(row = row, column = 1).value)[ :3].lower().find('bus') != -1:

                #Delete the row
                self.deleteRows(row, 1)
                #Decrement the counter
                row -= 1
                
            #Increment the counter
            row += 1

#Class that generates a MOM Report
class SwitchReport(ReportsTableHandler):

    #Function responsible for generating a single INT Report from a directory path
    #Using Regex to filter out the files we do not need
    #Make sure that the file naming convention is followed as stated in the User Guide
    
    def generateTable(self, dirPath, outputDir = os.getcwd(),
                      outputName = "Switch_Report.xlsx", regexName = "SWT_.*_.*_.*"):

        #------------------------INITIALIZING DEFAULT VARIABLES--------------------------#
        
        #Set the length of header columns and body columns
        hNumCols = 3
        bNumCols = 1

        sRow = 2
        sCol = "^Mom. Asymm. kA$"

        #Create a file names list for files in the specified directory
        fileNames = self.createFileList(dirPath, regexName)

        if len(fileNames) == 0:
            return

        #------------------------------CREATING HEADERS---------------------------------#
        
        #Specify the order of data columns with file names
        order = ["PRES", "GEN", "ULT", "CTT"]

        #All unordered header list generated from the filename **_TYPE_**
        unorderedHeaders = self.generateHeaders(dirPath, regexName)

        #Orderable ordered header list generated from fixed order list above
        orderedHeaders = self.listUnion(order, unorderedHeaders)

        #Remainder headers whose order does not matter
        remHeaders = self.listDiff(unorderedHeaders, orderedHeaders)

        #All ordered header list
        allOrderedHeaders = orderedHeaders + remHeaders

        #Reorder file names by using the order list
        fileNamesOrdered = self.fileListReorder(fileNames, order)

        #-------------------------------CREATING BODY-----------------------------------#
        
        #Loop through all the files in the reordered file list and generate excel files with its data
        self.loopData(dirPath, fileNamesOrdered, 3, 1, sRow, sCol, hNumCols,
                      bNumCols, addClosingCols = True)

        #-----------------------------FORMATTING TABLE---------------------------------#
        
        #Format the sheet
        self.formatTable(hNumCols, bNumCols)

        #------------------------------ADDING HEADERS--------------------------------#
        
        #Create a headers and subheaders list
        headers = ["Device"] + ["Momentary Duty: " + s for s in allOrderedHeaders] + ["Device Capability"]
        subheaders = ["ID", "Voltage", "Device"] + ["Mom. Asymm. kA"] * int(((self.ws.max_column - hNumCols)/bNumCols) - 1) + ["Rated Mom. Asymm. kA"]

        #Set the length of the merged header cells for the header columns and the body columns
        hMergeLen = hNumCols
        bMergeLen = bNumCols

        #Merge header cells and add headers
        self.mergeHeaderCells(hMergeLen, bMergeLen)
        self.addHeaders(hNumCols, bNumCols, bMergeLen , headers, subheaders)

        # Create a 2D list of columns to compare
        # The reference column comes first
        cmpCols = [["Rating Int. kA", item] for item in allOrderedHeaders]

        # Color fill the cells that do not meet the conditions in the function
        self.fillCmpCells(2, cmpCols)

        #---------------------------------SAVING------------------------------------#
        
        #Change the filename to some tag
        self.saveWorkbook(dirPath = outputDir, fileName = outputName)

        #Autofit columns
        self.autofitColumns(self.wbPath)

        
#Class that generates an Arc Flash Report
class ArcFlashReport(ReportsTableHandler):

    def generateTable(self, filePath, minCalorie = None, outputDir = os.getcwd()):

        #------------------------INITIALIZING DEFAULT VARIABLES--------------------------#
        
        #Set the length of header columns and body columns
        hNumCols = 20
        bNumCols = 0

        #---------------------------CREATING BODY & HEADERS------------------------------#
        
        #Get the file name into a list because the loopData function takes a list input
        fileNames = [Path(filePath).name]

        #If no file exists, exit function
        if len(fileNames) == 0:
            return
        
        #Loop through the single file in the file list and generate excel file with its data
        self.loopData(Path(filePath).parent, fileNames, 1, 1, 1, 1, hNumCols, bNumCols, numRows = True)

        #---------------------------------SAVING------------------------------------#

        #Close the workbook so that pywin32 can open it
        self.saveWorkbook(dirPath = outputDir, fileName = fileNames[0].split('.')[0] + '.xlsx')

        #---------------------------------SORTING------------------------------------#
        #Get the column with "Total Energy" in its heading
        totalEnergyCol = self.getColumnByHeaderName(self.ws, "Total Energy \(cal/cm²\)")

        #Sort the data by the "Total Energy" column
        self.sortByColumn(totalEnergyCol)

        #--------------------------------FORMATTING----------------------------------#
        
        #Reinitialize the class for reloading the sorted excel file
        self.__init__(self.wbPath)

        #Format the sheet
        self.formatTable(hNumCols, bNumCols)

        #Reduce the row size of subheaders row since we do not have subheaders in this case
        self.ws.row_dimensions[2].height = 15

        #Remove formatting from the subheader row since we do not have subheaders in this case
        for row in self.ws.iter_rows(min_row=2, max_col=self.ws.max_column, max_row=2):
            for cell in row:
                cell.font = Font(name='Calibri', size=9, color='FF000000')
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                self.ws.column_dimensions[get_column_letter(cell.column)].width = 20
                
        #If the "Total Energy" cells exceed minCalorie, fill them with red colour 
        self.checkEnergy(totalEnergyCol, minCalorie)

        #---------------------------------SAVING------------------------------------#
        
        #Save the file with .xlsx format
        self.saveWorkbook(dirPath = outputDir, fileName = fileNames[0].split('.')[0] + '.xlsx')

        #Autofit columns
        self.autofitColumns(self.wbPath)
        
    #---------------------------------UTIL FUNCTIONS------------------------------------#
        
    #Fills any cells in the "Total Energy" column with red colour if its value exceeds minVal
    def checkEnergy(self, col, minVal):

        #Loop through all the rows in the current sheet
        for row in range(self.ws.max_row):

            # Try for all cells. Cells that are empty will be skipped.
            try:
                if self.ws.cell(row = row + 1, column = col).value > minVal:
                    self.ws.cell(row = row + 1, column = col).fill = PatternFill("solid", fgColor="FF0000")
            except:
                pass

    #Uses pywin32 to sort the data by the "Total Energy" column
    def sortByColumn(self, col):

        # Initialize the excel COM
        excel = win32.gencache.EnsureDispatch("Excel.Application")

        # Open the workbook
        wb = excel.Workbooks.Open(self.wbPath)
        ws = wb.Worksheets(self.ws.title)

        # Sort by the column number col
        ws.Columns('A:Z').Sort(Key1=ws.Range(get_column_letter(col) + '1'),
                               Order1=2, Orientation=1, Header=1)

        # Save workbook
        wb.Save()
        excel.Application.Quit()
        del excel
