#------------------------------------------------------------------------------------------#
# Author: Adil Zafar Khan
# Last Edit Date: 12/22/2021
# Description:
"""
    ClassBuilder is a class that contains all the module functions that execute the program.
    It can be viewed as the program manager and every class is finally initialized and
    called from here.
"""
#------------------------------------------------------------------------------------------#

#Import required libraries
import os
import re
import pythoncom
from copy import copy
from pathlib import Path
import win32com.client as win32
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side

#Parent class for generating excel reports
class ReportsTableHandler:

    #------------------------INITIALIZING FUNCTIONS--------------------------#

    #Initialize class
    def __init__(self, wbPath = None):

        if wbPath == None:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Reports Table"
        else:
            self.wb = load_workbook(wbPath)
            self.ws = self.wb.active

    #Add a new sheet to the workbook
    def addSheet(self, sheetName = "New_Sheet"):
        newSheet = self.wb.create_sheet(sheetName)
        self.ws = newSheet
        return newSheet

    #Save workbook at the specified path
    def saveWorkbook(self, dirPath = Path.cwd(), fileName = "Reports_Table.xlsx"):
        self.wb.save(Path(dirPath, fileName))
        self.wbPath = Path(dirPath, fileName)

    #Function that converts an xlx file to xlsx
    def convertToXlsx(self, path):

        pythoncom.CoInitialize()
        
        if path.suffix == '.xlsx':
            return path

        #Open excel in pywin32
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(path)
        
        #FileFormat = 51 is for .xlsx extension
        wb.SaveAs(str(path) + "x", FileFormat = 51)    
        wb.Close()
        excel.Application.Quit()
        del excel
        
        return Path(str(path) + "x")

    #Creates an empty excel file at the pecified path
    def createEmptyExcelFile(self, wbPath):

        newWb = Workbook()
        newWb.save(wbPath)

    #------------------------FILE HANDLING FUNCTIONS--------------------------#
        
    #Generates a set of files that contain unique strings from file name
    def createFileSet(self, dirPath, regexStr = "(?<=_).*?(?=_)", regexIndex = 0):

        fileSet = set()

        for fileName in os.listdir(dirPath):
            if fileName.endswith(".xls"):
                fileStem = fileName.split('.')[0]
                fileSet.add(re.findall(regexStr, fileName)[regexIndex])

        return list(fileSet)

    #Generates a list that is the intersection of list1 and list2 with
    #the order of list1 preserved in the new list
    def listUnion(self, list1, list2):

        union = []
        for item in list1:
            for subitem in list2:
                if item == subitem:
                    union.append(item)
                    
        return union

    #Generates a complement of intersection list, list1 - list2
    def listDiff(self, list1, list2):
        return list(set(list1) - set(list2)) + list(set(list2) - set(list1))
    
    #Creates a file list that contains the names of all required xlsx files
    def createFileList(self, dirPath, regexName = None):

        #Group is the last letters of a filename
        fileNames = []
        #Loop through directory
        for fileName in os.listdir(dirPath):

            #If group is NOT passed and the file has an xls extension
            if fileName.endswith(".xls") and regexName == None:
                
                #Append the file name to the fileNames list
                fileNames.append(fileName)
                
            else:

                #If the file has an xls extension and the regexName matches the fileName
                if fileName.endswith(".xls") and re.search(regexName, fileName.split('.')[0]) != None:

                    #Append the file name to the fileNames list
                    fileNames.append(fileName)

        return fileNames

    def fileListReorder(self, fileNames, order, last = None):

        #It does not matter if the items in order list repeat. As long as the items do not repeat in both fileNames and order
        #the final list will have unique elements

        #Create an empty final list
        orderedList = []
        
        #Flag that checks if we have moved the last file to the end of the fileNames list
        lastSet = False
        
        #Loop through each item in order list
        for item in order:
            i = 0
            
            #Loop through each item in fileNames list
            while i < len(fileNames):
                
                #If the filename at index i contains item, then remove it from fileNames and append it to the orderedList
                if fileNames[i].find(item) != -1:
                    orderedList.append(fileNames.pop(i))
                    
                #Else if the filename at index i contains last, then move it to the end of the fileNames
                #list, set lastSet = True and decrement i so
                #the loop does not skip the item that descends into the last item index in fileNames
                #elif fileNames[i].find(last) != -1 and lastSet == False:
                    
                    #fileNames.append(fileNames.pop(i))
                    #lastSet = True
                    #i = i - 1
                    
                i += 1
                
        #Append the remaining fileNames list to the orderedList list         
        orderedList += fileNames
        return orderedList

    def createFillerFiles(self, dirPath, regexName = None):

        typeCountDict = {}
        counterSet = set()
        xlsList = self.createFileList(dirPath, regexName)

        for file in xlsList:
            
            fileType = file.split('_')[3]
            typeCountDict[fileType] = typeCountDict.get(fileType, 0) + 1

        try:            
            max_key = max(typeCountDict, key=typeCountDict.get)
            maxKeyList = self.createFileList(dirPath, regexName = regexName.replace(".*", max_key, 1))
            maxGroupList =  [item[item.rfind('_') + 1 : ] for item in maxKeyList]
        except:
            pass
        
        xlsListString = ''
        for item in xlsList:
            xlsListString += item
        
        for key in typeCountDict:
            for item in maxGroupList:
                regexStr = regexName.replace(".*", key, 1).replace(".*", item, 1)
                if re.search(regexStr, xlsListString) == None:
                    self.createEmptyExcelFile(Path(dirPath, regexStr))

    #------------------------WORKSHEET HANDLING FUNCTIONS--------------------------#
                    
    #Merges header cells.
    #hMergeLen = the number of the first x columns
    #bMergeLen = the number of the first y columns after the x columns
    
    def mergeHeaderCells(self, hMergeLen, bMergeLen):

        i = 0
        #Loop through all the columns
        while i < self.ws.max_column - 1:

            #If it is the first iteration
            if i == 0:

                #Merge the header cells for the header columns
                self.ws.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = hMergeLen)
                i += hMergeLen + 1
                
            else:

                #Merge the header cells for the body columns
                self.ws.merge_cells(start_row = 1, start_column = i + 1, end_row = 1, end_column = i + bMergeLen)
                i += bMergeLen + 1

    def generateHeaders(self, dirPath, regexName = ".*_.*_.*_.*"):

        headerSet = set()
        
        for fileName in os.listdir(dirPath):
            if fileName.endswith(".xls") and re.search(regexName, fileName) != None:

                headerSet.add(re.findall("(?<=_).*?(?=_)", fileName)[-1])

        return list(headerSet)

                
    #Add headers to the table
    def addHeaders(self, hNumCols, bNumCols, bMergeLen, headerList = [], subheaderList = []):

        #headerList counter, starts from zero so we will add 1 to it since openpyxl starts with 1
        i = 0

        #Adding Headers (Row = 1)
        for item in headerList:

            #If it is the first item
            if headerList.index(item) == 0:

                #It is a header for the header columns
                self.ws.cell(row = 1, column = i + 1).value = item
                i += hNumCols + 1
                    
            else:

                #It is a header for the body columns
                self.ws.cell(row = 1, column = i + 1).value = item

                i += bMergeLen + 1

            #If the headerList counter exceeds max number of filled columns, exit out of the loop
            if i >= self.ws.max_column:
                break

        #subheaderList counter, starts from zero so we will add 1 to it since openpyxl starts with 1
        i = 0

        #Adding Subheaders (Row = 2)
        for item in subheaderList:

            #Skip if the cell is an empty cell (width of the empty cells has been set to 4)
            if self.ws.column_dimensions[get_column_letter(i + 1)].width == 4:
                i += 1

            #Fill in the subheader cell
            self.ws.cell(row = 2, column = i + 1).value = item
            i += 1

            #If the subheaderList counter exceeds max number of filled columns, exit out of the loop
            if i >= self.ws.max_column:
                break

    def getColumnByHeaderName(self, sheet, name, headerRowNum = 1):
        
        for col in range(1, sheet.max_column + 1):
            
            if re.search(name, str(sheet.cell(row = headerRowNum, column = col).value)) != None:
                return col

        return 1

    def matchIDs(self, ID, sRow):
        for i in range(sRow, self.ws.max_row + 1):
            if self.ws.cell(row = i, column = 1).value == ID:
                return i

        return self.ws.max_row + 1

    def transferColumnsPreserved(self, sheet, dRow, dCol, sRow, sCol, numCols, numRows = False):
        
        idRow = dRow
        #If numRows is not passed
        if not numRows:

            #Set the last row as the last filled row in the source excel sheet
            lastRow = sheet.max_row + 1
            
        else:
            
            #Set the last row as the position of the first empty row
            lastRow = self.getFirstEmptyRowNum(sheet) + 1

        #Last column position = starting column + number of columns
        lastCol = sCol + numCols
        
        #Counters for counting through cells
        counterCol = 0
        
        #Loop through each cell in the range of start row, start column and last row, last column of the source sheet
        for row in range(sRow, lastRow):
            for col in range(sCol, lastCol):

                dRow = self.matchIDs(sheet.cell(row = row, column = 1).value, idRow)
                #Copy and paste the content in the cell to the counters + starting position 
                self.ws.cell(row = dRow, column = dCol + counterCol).value = sheet.cell(row = row, column = col).value

                #Convert to numbers
                try:

                    #If it is a number
                    if (self.ws.cell(row = dRow, column = dCol + counterCol).value.replace('.','').isnumeric()):
                        
                        #Convert it to float and round it to two decimal places
                        numericVal = float(self.ws.cell(row = dRow, column = dCol + counterCol).value)
                        self.ws.cell(row = dRow, column = dCol + counterCol).value = numericVal
                except:
                    pass
                
                #Increment column counter
                counterCol += 1
                
            #Reset column counter
            counterCol = 0

    def transferColumns(self, sheet, dRow, dCol, sRow, sCol, numCols, numRows = False):

        #If numRows is not passed
        if not numRows:

            #Set the last row as the last filled row in the source excel sheet
            lastRow = sheet.max_row + 1
            
        else:
            
            #Set the last row as the position of the first empty row
            lastRow = self.getFirstEmptyRowNum(sheet) + 1

        #Last column position = starting column + number of columns
        lastCol = sCol + numCols
        
        #Counters for counting through cells
        counterCol = 0
        counterRow = 0
        
        #Loop through each cell in the range of start row, start column and last row, last column of the source sheet
        for row in range(sRow, lastRow):
            for col in range(sCol, lastCol):

                #Copy and paste the content in the cell to the counters + starting position 
                self.ws.cell(row = dRow + counterRow, column = dCol + counterCol).value = sheet.cell(row = row, column = col).value

                #Convert to numbers
                try:

                    #If it is a number
                    if (self.ws.cell(row = dRow + counterRow, column = dCol + counterCol).value.replace('.','').isnumeric()):
                        
                        #Convert it to float and round it to two decimal places
                        numericVal = float(self.ws.cell(row = dRow + counterRow, column = dCol + counterCol).value)
                        self.ws.cell(row = dRow + counterRow, column = dCol + counterCol).value = numericVal
                        
                except:
                    pass

                #Increment column counter
                counterCol += 1
                
            #Reset column counter
            counterCol = 0
            
            #Increment row counter
            counterRow += 1


    #Loops each excel file in a directory, extracts data from it and places it in the worksheet object of the class
    def loopData(self, dirPath, fileNames, dRow, dCol, sRow, sCol, hNumCols, bNumCols, numRows = False,
                 addClosingCols = False, numFiles = None):

        #dRow = Destination start row
        #dCol = Destination start col
        #sRow = Source start row
        #sRow = Source start col
        #hNumCols = Number of header cols
        #bNumCols = Number of body cols
        #numRows = Flag indicating to stop at the first empty row
        #addClosingCols = Flag indicating to add a final closing column
        #numFiles = Number of files to iterate through

        # if destination start row is already filled, start from the first empty row
        dRow = self.ws.max_row - 1 + dRow
        
        # Save the position of the destination starting column, used for header columns
        idCol = dCol

        #Jump over the header columns
        dCol += hNumCols + 1

        #Checks if header columns hav been transferred
        hColsTransferred = False

        #If numFiles is not passed, loop through all the files in the fileNames list
        if numFiles == None:
            numFiles = len(fileNames)

        #Loop through the fileNames list
        for fileName in fileNames:

            #Set the original file path
            OrgFilePath = Path(dirPath, fileName)

            #Convert xls to xlsx
            filePath = self.convertToXlsx(OrgFilePath)

            #Load the excel file using openpyxl library
            dataWb = load_workbook(filePath)
            dataSheet = dataWb.worksheets[0]

            #check if sCol is a string passed, change it to a number
            #Only run it once on the first iteration
            if isinstance(sCol, str):
                sColNum = self.getColumnByHeaderName(dataSheet, sCol)
            else:
                sColNum = sCol

            #If it is the first column and the header columns exist
            if not hColsTransferred and hNumCols > 0 and dataSheet.max_column > 1:
                    
                #Transfer the header columns
                self.transferColumns(dataSheet, dRow, idCol, sRow, 1, hNumCols, numRows = numRows)
                hColsTransferred = True
                
            else:
                self.transferColumnsPreserved(dataSheet, dRow, idCol, sRow, 1, hNumCols, numRows = numRows)
                
            #Transfer the body columns
            self.transferColumnsPreserved(dataSheet, dRow, dCol, sRow, sColNum, bNumCols, numRows = numRows)

            #If addClosingCols is True
            if addClosingCols:

                startLastCol = len(fileNames) * bNumCols + hNumCols + len(fileNames) + 2

                #Transfer an extra column to the right of the last column
                self.transferColumnsPreserved(dataSheet, dRow, startLastCol, sRow, hNumCols + 1, bNumCols, numRows = numRows)

            #Close the excel workbook
            dataWb.close()

            #Jump to the next insertion location of the body columns
            dCol += bNumCols + 1

            #Delete the extra xlsx file created if the original file was xls
            if Path(OrgFilePath).suffix == '.xls':
                os.remove(filePath)

            #If numFiles is passed then only iterate through the number of files specified in numFiles
            if fileNames.index(fileName) >= numFiles - 1:
                break

    #Formats the excel sheet
    def formatTable(self, hNumCols, bNumCols):

        #Set the font for the body rows
        ft2 = Font(name='Calibri', size=9, vertAlign=None, color='FF000000')

        #Set the font for the header rows
        ft1 = copy(ft2)
        ft1.bold = True

        #Set the colour for the heading rows
        colorHeading = PatternFill("solid", fgColor="BFBFBF")

        #Set the colour for the even body rows
        colorEven = PatternFill("solid", fgColor="D9D9D9")

        #Set the border for the cells
        thin = Side(border_style="thin", color="000000")

##        #Increase the height of the header rows
##        self.ws.row_dimensions[1].height = 20
##        self.ws.row_dimensions[2].height = 20

        #Initialize a counter for rows
        rowNum = 1

        #Set the column width to 14 points for the columns that contain data
        for i in range(1, self.ws.max_column + 1, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = 14
            self.ws.cell(1, i).border = Border(top = thin, bottom = thin)

        #Set the column width to 4 points for the columns that are empty
        for i in range(hNumCols + 1, self.ws.max_column + 1, bNumCols + 1):
            self.ws.column_dimensions[get_column_letter(i)].width = 4

        #Loop through each cell in the table
        for row in self.ws.iter_rows(min_row=1, max_col=self.ws.max_column, max_row=self.ws.max_row):
            for cell in row:

                #If it is a header row   
                if rowNum < 3:

                    #Format header rows
                    cell.font = ft1
                    cell.fill = colorHeading

                else:

                    #Format body rows
                    cell.font = ft2
                    
                    #If it is an even body row
                    if rowNum%2 == 0:

                        #Fill the cells with a colour
                        cell.fill = colorEven
                        
                #If the cell is a data cell
                if cell.value != None or self.ws.column_dimensions[get_column_letter(cell.column)].width > 5:

                    #Add a border
                    cell.border = Border(top=thin, left=thin, bottom = thin, right=thin)
    
                elif rowNum > 1:

                    #Fill in with white color
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")

                #Centre align all cells
                cell.alignment = Alignment(horizontal="center", vertical="center")

            #Increment row counter
            rowNum += 1

    #Auto fit the columns of a given excel sheet
    def autofitColumns(self, wbPath):

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(wbPath)
        ws = wb.Worksheets(1)

        ws.UsedRange.EntireColumn.NumberFormat = "0.00"
        
        #Auto fit used columns
        ws.UsedRange.EntireColumn.AutoFit()

        for i in range(1, self.ws.max_column + 1):

            for j in range(2, self.ws.max_row + 1):

                if isinstance(self.ws.cell(j, i).value, int) or str(self.ws.cell(j, i).value).endswith(".0"):
                    ws.Cells(j, i).NumberFormat = "0"

            if self.ws.cell(1, i).value == "Voltage" or self.ws.cell(1, i).value == "kV":

                for j in range(2, self.ws.max_row + 1):
                    
                    if isinstance(self.ws.cell(j, i).value, float):
                        ws.Cells(j, i).NumberFormat = "0.000"
                            
            if self.ws.cell(2, i).value == "Voltage" or self.ws.cell(2, i).value == "kV":

                for j in range(3, self.ws.max_row + 1):
                    
                    if isinstance(self.ws.cell(j, i).value, float):
                        ws.Cells(j, i).NumberFormat = "0.000"

        #Auto fit used rows
        ws.UsedRange.EntireRow.AutoFit()

        #Close and save
        wb.Close(SaveChanges = True)
        del excel

    #Moves columns without overwriting existing columns at the destination
    def moveColumn(self, colNum, xMove):

        #Insert an empty column to the right
        if xMove > 0:
            self.ws.insert_cols(colNum + xMove - 1)

        #Insert an empty column to the left
        else:
            self.ws.insert_cols(colNum + xMove + 1)

        #Get the column letter from its position
        col = get_column_letter(colNum + 1)
        
        #Move the column contents and overwrite the empty column
        self.ws.move_range(col + '1:' + col + str(self.ws.max_row), cols=xMove)

        #Delete the old column
        self.ws.delete_cols(colNum + 1)

    #Deletes the specified number of columns
    def deleteColumns(self, startCol, numCols):

        self.ws.delete_cols(startCol, numCols)

    def deleteRows(self, startRow, numRows):

        self.ws.delete_rows(startRow, numRows)

    #Returns the position of the first row that is empty in an excel sheet
    def getFirstEmptyRowNum(self, sheet):

        #Loop through the first cell in each row
        for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value == None:
                    return cell.row

        return sheet.max_row

    #Fills any cells in the passed second col of cmpCols list with red colour if its value exceeds
    #the value of the first col of cmpCols
    def fillCmpCells(self, headerRowNum, cmpCols):

        for i in range(len(cmpCols)):
            
            if isinstance(cmpCols[i][0], str):
                cmpCols[i][0] = self.getColumnByHeaderName(self.ws, cmpCols[i][0], headerRowNum)

            if isinstance(cmpCols[i][1], str):
                cmpCols[i][1] = self.getColumnByHeaderName(self.ws, cmpCols[i][1], headerRowNum)
            
        for row in range(headerRowNum + 1, self.ws.max_row):
            for item in cmpCols:
                
                try:
                    float(self.ws.cell(row = row, column = item[0]).value)
                    float(self.ws.cell(row = row, column = item[1]).value)
                    
                    if self.ws.cell(row = row, column = item[1]).value > self.ws.cell(row = row, column = item[0]).value:
                        self.ws.cell(row = row, column = item[1]).fill = PatternFill("solid", fgColor="FF0000")
                except:
                    pass


    #------------------------EXECUTION FUNCTION (MOM & INT)-------------------------#
                
    #Builds all the required tables. Only used for INT and MOM reports.
    def buildTables(self, inDirPath, outDirPath, form):
        
        fileSetForm = self.createFileSet(inDirPath, regexStr = "(?<=_).*?(?=_)", regexIndex = 0)
        fileSetType = self.createFileSet(inDirPath, regexStr = "(?<=_).*?(?=_)", regexIndex = 1)

        for name in fileSetForm:
            for subname in fileSetType:
                
                regexName = form + "_" + name + '_' + subname + "_.*"
                self.__init__()
                self.createFillerFiles(inDirPath, regexName + "_.*")
                self.generateTable(inDirPath, outDirPath, regexName.replace("_.*", ".xlsx"), regexName)

